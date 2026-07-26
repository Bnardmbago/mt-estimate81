from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.exports.markdown import format_currency
from app.exports.theme import BLUE_LIGHT, BLUE_PRIMARY, TEXT_ON_PRIMARY, YELLOW_SECTION, YELLOW_TOTAL
from app.models.estimate import Estimate

# Branded color output is PDF/XLSX only; Markdown exports remain plain text.

SHEET_NAMES = {
    "en": {
        "executive": "Executive",
        "features": "Features",
        "timeline": "Timeline",
        "nrc": "NRC Detail",
        "rc": "RC Detail",
        "assumptions": "Assumptions",
        "reference": "Reference",
    },
    "ja": {
        "executive": "エグゼクティブ",
        "features": "機能詳細",
        "timeline": "タイムライン",
        "nrc": "NRC内訳",
        "rc": "RC内訳",
        "assumptions": "前提条件",
        "reference": "参照",
    },
}

CURRENCY_FORMAT = '"¥"#,##0'


def _solid_fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _table_header_font() -> Font:
    return Font(bold=True, color=TEXT_ON_PRIMARY)


def _section_font() -> Font:
    return Font(bold=True)


def _total_font() -> Font:
    return Font(bold=True)


def _apply_table_header(cell) -> None:
    cell.fill = _solid_fill(BLUE_PRIMARY)
    cell.font = _table_header_font()


def _apply_section_title(cell) -> None:
    cell.fill = _solid_fill(YELLOW_SECTION)
    cell.font = _section_font()


def _apply_label_cell(cell) -> None:
    cell.fill = _solid_fill(YELLOW_SECTION)
    cell.font = _section_font()


def _apply_total_row(cells: list) -> None:
    fill = _solid_fill(YELLOW_TOTAL)
    font = _total_font()
    for cell in cells:
        cell.fill = fill
        cell.font = font


def _apply_zebra_row(cells: list, *, data_row_index: int) -> None:
    if data_row_index % 2 == 0:
        fill = _solid_fill(BLUE_LIGHT)
        for cell in cells:
            cell.fill = fill


def _header_font() -> Font:
    return _section_font()


def _write_key_value_rows(
    ws,
    rows: list[tuple[str, Any]],
    *,
    start_row: int = 1,
    value_format: str | None = None,
) -> int:
    row_idx = start_row
    for label, value in rows:
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        _apply_label_cell(label_cell)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if value_format:
            cell.number_format = value_format
        row_idx += 1
    return row_idx


def _write_section_title(ws, row: int, column: int, value: str) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    _apply_section_title(cell)


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    start_row: int = 1,
    column_formats: dict[int, str] | None = None,
) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        _apply_table_header(cell)

    row_idx = start_row + 1
    for data_index, row in enumerate(rows):
        row_cells = []
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if column_formats and col_idx in column_formats:
                cell.number_format = column_formats[col_idx]
            row_cells.append(cell)
        _apply_zebra_row(row_cells, data_row_index=data_index)
        row_idx += 1
    return row_idx


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def _write_bullet_list(ws, items: list[str], *, start_row: int) -> int:
    row_idx = start_row
    for item in items:
        ws.cell(row=row_idx, column=1, value=f"- {item}")
        row_idx += 1
    return row_idx


def _build_executive_pricing_rows(ctx: dict[str, Any]) -> list[tuple[str, Any]]:
    labels = ctx["labels"]
    pricing = ctx.get("pricing_summary") or {}
    if pricing.get("has_discount"):
        return [
            (
                labels["development_cost_original"],
                format_currency(pricing["nrc_original_total_jpy"]),
            ),
            (labels["limited_time_discount"], pricing["discount_display"]),
            (
                labels["special_price"],
                f"{format_currency(pricing['nrc_discounted_total_jpy'])} {labels['excluding_tax']}",
            ),
        ]
    return [
        (
            labels["total_development_cost"],
            format_currency(ctx["executive_display"]["development_cost_jpy"]),
        ),
    ]


def _build_executive_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    project = ctx["project_summary"]
    executive_display = ctx["executive_display"]

    row_idx = _write_key_value_rows(
        ws,
        [
            (labels["project_name"], project["project_name"]),
            (labels["estimate_type"], project["estimate_type"]),
            (labels["client_name"], project["client_name"]),
            (labels["estimate_id"], project["estimate_id"]),
            (labels["export_revision"], project["export_revision"]),
            (labels["generated_date"], project["generated_date"]),
            (labels["estimate_creator"], project["estimate_creator"]),
        ],
    )
    row_idx += 1

    _write_section_title(ws, row_idx, 1, labels["executive_cost_summary"])
    row_idx += 1
    summary_rows = [
        *_build_executive_pricing_rows(ctx),
        (
            labels["maintenance_cost_monthly_annual"],
            executive_display["maintenance_cost_display"],
        ),
        (
            labels["development_period"],
            executive_display["development_period_display"],
        ),
    ]
    for label, value in summary_rows:
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        _apply_label_cell(label_cell)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    pricing = ctx.get("pricing_summary") or {}
    if pricing.get("has_discount") and pricing.get("campaign_terms"):
        row_idx += 1
        _write_section_title(ws, row_idx, 1, pricing.get("campaign_terms_title", labels["campaign_terms_title"]))
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=pricing["campaign_terms"])
        row_idx += 1

    _write_section_title(ws, row_idx, 1, labels["questionnaire"])
    row_idx += 1
    questionnaire_sections = ctx.get("questionnaire_sections") or []
    if questionnaire_sections:
        for section in questionnaire_sections:
            _write_section_title(ws, row_idx, 1, section["title"])
            row_idx += 1
            row_idx = _write_key_value_rows(
                ws,
                [(field["label"], field["value"]) for field in section["fields"]],
                start_row=row_idx,
            )
            row_idx += 1
    else:
        ws.cell(row=row_idx, column=1, value=labels["none"])
        row_idx += 1


def _build_features_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    _write_table(
        ws,
        [
            labels["feature_name"],
            labels["feature_description"],
            labels["feature_phase"],
            labels["feature_role"],
            labels["feature_hours"],
            labels["feature_days"],
        ],
        [
            [
                item["name"],
                item["description"],
                item["phase"],
                item["role"],
                item["hours"],
                item["days"],
            ]
            for item in ctx.get("feature_items") or []
        ],
    )


def _build_timeline_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    gantt = ctx.get("gantt") or {}
    tasks = gantt.get("tasks") or []

    if not tasks:
        ws.cell(row=1, column=1, value=labels["none"])
        return

    _write_key_value_rows(
        ws,
        [
            (labels["gantt_project_start"], gantt.get("project_start_date")),
            (labels["gantt_project_end"], gantt.get("project_end_date")),
            (labels["gantt_total_working_days"], gantt.get("total_working_days")),
        ],
    )


def _build_nrc_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    calculation = ctx["calculation"]
    line_items = calculation.get("nrc_line_items") or []
    row_idx = _write_table(
        ws,
        [labels["category"], labels["item"], labels["cost"]],
        [[row["category"], row["item"], row["cost_jpy"]] for row in line_items],
        column_formats={3: CURRENCY_FORMAT},
    )
    row_idx += 1
    total_label = ws.cell(row=row_idx, column=1, value=labels["nrc_total"])
    total_cell = ws.cell(row=row_idx, column=3, value=ctx["executive_summary"]["nrc_total_jpy"])
    total_cell.number_format = CURRENCY_FORMAT
    _apply_total_row([total_label, total_cell])


def _build_rc_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    rc_breakdown = ctx.get("rc_breakdown") or {}
    line_items = rc_breakdown.get("line_items") or []
    rows = []
    for row in line_items:
        item_label = row.get("service_description") or (
            labels["maintenance"] if row.get("is_maintenance") else row["item"]
        )
        rows.append(
            [
                row["category"],
                item_label,
                row["monthly_jpy"],
                row["annual_jpy"],
            ]
        )
    row_idx = _write_table(
        ws,
        [labels["category"], labels["service_description"], labels["monthly"], labels["annual"]],
        rows,
        column_formats={3: CURRENCY_FORMAT, 4: CURRENCY_FORMAT},
    )
    row_idx += 1
    monthly_total_label = ws.cell(row=row_idx, column=1, value=labels["monthly_total"])
    monthly_total_cell = ws.cell(
        row=row_idx,
        column=3,
        value=rc_breakdown.get("monthly_total_jpy", 0),
    )
    monthly_total_cell.number_format = CURRENCY_FORMAT
    _apply_total_row([monthly_total_label, monthly_total_cell])
    row_idx += 1
    annual_total_label = ws.cell(row=row_idx, column=1, value=labels["annual_total"])
    annual_total_cell = ws.cell(
        row=row_idx,
        column=4,
        value=rc_breakdown.get("annual_total_jpy", 0),
    )
    annual_total_cell.number_format = CURRENCY_FORMAT
    _apply_total_row([annual_total_label, annual_total_cell])


def _build_assumptions_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    extracted = ctx["extracted"]
    row_idx = 1

    _write_section_title(ws, row_idx, 1, labels["questionnaire"])
    row_idx += 1
    questionnaire_sections = ctx.get("questionnaire_sections") or []
    if questionnaire_sections:
        for section in questionnaire_sections:
            _write_section_title(ws, row_idx, 1, section["title"])
            row_idx += 1
            for field in section["fields"]:
                label_cell = ws.cell(row=row_idx, column=1, value=field["label"])
                _apply_label_cell(label_cell)
                ws.cell(row=row_idx, column=2, value=field["value"])
                row_idx += 1
            row_idx += 1
    else:
        ws.cell(row=row_idx, column=1, value=labels["none"])
        row_idx += 1
    row_idx += 1

    sections = [
        (labels["functional_requirements"], extracted.get("functional_requirements")),
        (labels["non_functional_requirements"], extracted.get("non_functional_requirements")),
        (labels["modules"], extracted.get("modules")),
        (labels["user_roles"], extracted.get("user_roles")),
        (labels["external_systems"], extracted.get("external_systems")),
    ]
    has_content = False
    for section_label, items in sections:
        if not items:
            continue
        has_content = True
        _write_section_title(ws, row_idx, 1, section_label)
        row_idx += 1
        row_idx = _write_bullet_list(ws, items, start_row=row_idx)
    if not has_content:
        _write_section_title(ws, row_idx, 1, labels["functional_requirements"])
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=labels["none"])


def _build_reference_sheet(ws, ctx: dict[str, Any]) -> None:
    labels = ctx["labels"]
    extracted = ctx["extracted"]
    row_idx = 1

    _write_section_title(ws, row_idx, 1, labels["estimate_exclusions"])
    row_idx += 1
    exclusions = extracted.get("estimate_exclusions") or []
    if exclusions:
        _write_bullet_list(ws, exclusions, start_row=row_idx)
    else:
        ws.cell(row=row_idx, column=1, value=labels["none"])


def add_report_sheets(wb: Workbook, report_context: dict[str, Any]) -> None:
    """Build the standard report sheets (Executive..Reference) into `wb`.

    Shared by `generate_excel` and the internal dossier XLSX generator so the
    report layout is defined in exactly one place.
    """
    locale = report_context["locale"]
    if locale not in ("ja", "en"):
        raise ValueError(f"Unsupported locale: {locale}")

    sheet_names = SHEET_NAMES[locale]
    builders = [
        (sheet_names["executive"], _build_executive_sheet),
        (sheet_names["features"], _build_features_sheet),
        (sheet_names["timeline"], _build_timeline_sheet),
        (sheet_names["nrc"], _build_nrc_sheet),
        (sheet_names["rc"], _build_rc_sheet),
        (sheet_names["assumptions"], _build_assumptions_sheet),
        (sheet_names["reference"], _build_reference_sheet),
    ]

    for name, builder in builders:
        ws = wb.create_sheet(name)
        builder(ws, report_context)
        _auto_width(ws)


def generate_excel(report_context: dict[str, Any], estimate: Estimate) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    add_report_sheets(wb, report_context)

    del estimate  # retained for API compatibility; workbook uses report_context only

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
