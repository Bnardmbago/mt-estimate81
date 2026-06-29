from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.exports.excel import SHEET_NAMES, generate_excel
from tests.unit.export_fixtures import sample_estimate_with_calculation, sample_report_context


def _load_workbook_from_bytes(content: bytes):
    return load_workbook(BytesIO(content), data_only=False)


@pytest.fixture
def report_context():
    return sample_report_context()


@pytest.fixture
def estimate():
    return sample_estimate_with_calculation()


def test_excel_workbook_has_seven_sheets(report_context, estimate):
    content = generate_excel(report_context, estimate)
    wb = _load_workbook_from_bytes(content)

    expected_sheets = list(SHEET_NAMES["en"].values())
    assert wb.sheetnames == expected_sheets
    assert len(wb.sheetnames) == 7


def test_excel_timeline_sheet_contains_gantt_summary(report_context, estimate):
    content = generate_excel(report_context, estimate)
    wb = _load_workbook_from_bytes(content)

    timeline_sheet = wb[SHEET_NAMES["en"]["timeline"]]
    assert timeline_sheet["A1"].value == "Project start"
    assert timeline_sheet["B1"].value == "2026-06-09"
    assert timeline_sheet["A3"].value == "Total working days"


def test_excel_nrc_total_on_nrc_detail_sheet(report_context, estimate):
    content = generate_excel(report_context, estimate)
    wb = _load_workbook_from_bytes(content)

    nrc_sheet = wb[SHEET_NAMES["en"]["nrc"]]
    last_row = nrc_sheet.max_row
    assert nrc_sheet.cell(row=last_row, column=1).value == "NRC Total"
    assert nrc_sheet.cell(row=last_row, column=3).value == 700000


def test_excel_executive_contains_project_name(report_context, estimate):
    content = generate_excel(report_context, estimate)
    wb = _load_workbook_from_bytes(content)

    executive_sheet = wb[SHEET_NAMES["en"]["executive"]]
    assert executive_sheet["B1"].value == "Portal Redesign"
    assert executive_sheet["B2"].value == "Web Application"
    assert executive_sheet["B3"].value == "ACME Corp"
    assert executive_sheet["B5"].value == 1


def test_excel_ja_locale_sheet_names(estimate):
    ctx = sample_report_context(estimate, locale="ja")
    content = generate_excel(ctx, estimate)
    wb = _load_workbook_from_bytes(content)

    assert wb.sheetnames == list(SHEET_NAMES["ja"].values())


def test_excel_rc_sheet_includes_monthly_and_annual_totals(report_context, estimate):
    content = generate_excel(report_context, estimate)
    wb = _load_workbook_from_bytes(content)

    rc_sheet = wb[SHEET_NAMES["en"]["rc"]]
    values = [
        (rc_sheet.cell(row=row, column=1).value, rc_sheet.cell(row=row, column=3).value, rc_sheet.cell(row=row, column=4).value)
        for row in range(1, rc_sheet.max_row + 1)
    ]
    assert ("Monthly RC Total", 170000, None) in values
    assert ("Annual RC Total", None, 2040000) in values
    assert ("Cloud Infrastructure", 50000, 600000) in values
    assert ("Maintenance and Support", 120000, 1440000) in values
