from io import BytesIO
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

from app.exports.excel import SHEET_NAMES, generate_excel
from app.exports.markdown import format_currency, format_effort_days, format_hours, format_person_days
from app.exports.theme import PRIMARY, SURFACE
from tests.unit.export_fixtures import (
    sample_estimate_with_calculation,
    sample_quotation_context,
    sample_report_context,
)

TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "app" / "exports" / "templates"


def _render_template(template_name: str, **context) -> str:
    env = Environment(
        loader=FileSystemLoader(TEMPLATE_DIR),
        autoescape=select_autoescape(["html", "xml"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )
    template = env.get_template(template_name)
    return template.render(**context)


def test_report_template_includes_readability_styles():
    html = _render_template(
        "estimate_report.html.j2",
        ctx=sample_report_context(),
        format_currency=format_currency,
        format_hours=format_hours,
        format_effort_days=format_effort_days,
        format_person_days=format_person_days,
    )
    assert "--export-border: #E2E8F0" in html
    assert "qa-table" in html
    assert "data-table" in html
    assert "gantt-chart" in html
    assert "margin: 14mm 12mm 16mm 12mm" in html


def test_quotation_template_includes_readability_styles():
    html = _render_template(
        "estimate_quotation.html.j2",
        ctx=sample_quotation_context(locale="ja"),
        format_currency=format_currency,
    )
    assert "--export-border: #E2E8F0" in html
    assert "下記の通りお見積りいたします。" in html
    assert "items-table" in html
    assert "col-item" in html
    assert "col-unit" in html
    assert "assets/BI_logo.svg" in html
    assert "〒103-0027" in html
    assert "アーバンネット日本橋二丁目ビル 10階" in html
    assert "TEL：03-6262-0742" in html
    assert "MAIL ：ai@beyondai.co.jp" in html
    assert "住信SBIネット銀行 法人第一支店（ 106） 普通口座 2112728" in html
    assert "payment-footer" in html
    assert "questionnaire-appendix" not in html
    assert "数量" not in html


def test_excel_table_headers_use_primary_theme():
    content = generate_excel(
        sample_report_context(),
        sample_estimate_with_calculation(),
    )
    wb = load_workbook(BytesIO(content), data_only=False)
    features_sheet = wb[SHEET_NAMES["en"]["features"]]
    header_fill = features_sheet["A1"].fill.fgColor.rgb
    assert header_fill in {PRIMARY, f"00{PRIMARY}"}


def test_excel_section_titles_use_surface_theme():
    content = generate_excel(
        sample_report_context(),
        sample_estimate_with_calculation(),
    )
    wb = load_workbook(BytesIO(content), data_only=False)
    executive_sheet = wb[SHEET_NAMES["en"]["executive"]]
    section_fill = executive_sheet["A8"].fill.fgColor.rgb
    assert section_fill in {SURFACE, f"00{SURFACE}"}
