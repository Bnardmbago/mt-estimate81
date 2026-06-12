from pathlib import Path

from docx import Document
from openpyxl import load_workbook


class ExtractionError(Exception):
    pass


def _validate_path(file_path: str) -> Path:
    path = Path(file_path).resolve()
    uploads_root = Path("/data/uploads").resolve()

    if not str(path).startswith(str(uploads_root)):
        raise ExtractionError("Invalid file path")

    if not path.is_file():
        raise ExtractionError(f"File not found: {file_path}")

    return path


def extract_to_markdown(file_path: str, file_type: str) -> dict:
    path = _validate_path(file_path)
    normalized_type = file_type.lower()

    if normalized_type == "pdf":
        return _extract_pdf(path)
    if normalized_type == "docx":
        return _extract_docx(path)
    if normalized_type == "xlsx":
        return _extract_xlsx(path)

    raise ExtractionError(f"Unsupported file type: {file_type}")


def _extract_pdf(path: Path) -> dict:
    import fitz

    document = fitz.open(path)
    try:
        pages = [page.get_text("text").strip() for page in document]
        markdown = "\n\n".join(text for text in pages if text)
        return {
            "markdown": markdown,
            "page_count": document.page_count,
            "method": "pymupdf",
        }
    finally:
        document.close()


def _extract_docx(path: Path) -> dict:
    document = Document(path)
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    markdown = "\n\n".join(paragraphs)
    return {
        "markdown": markdown,
        "page_count": None,
        "method": "python-docx",
    }


def _extract_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sections: list[str] = []
        for sheet in workbook.worksheets:
            sections.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    sections.append(" | ".join(cells))

        return {
            "markdown": "\n".join(sections),
            "page_count": len(workbook.worksheets),
            "method": "openpyxl",
        }
    finally:
        workbook.close()
